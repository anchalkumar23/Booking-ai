import csv
import io
import json
import re
import uuid
import openpyxl
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from sqlalchemy import func
from app.core.database import get_db
from app.api.v1.auth import _get_current_user
from app.models.promo_campaign import PromoCampaign, CampaignChannel
from app.models.location import Location
from app.models.whatsapp_message import WhatsAppMessage, WADirection, WAStatus
from app.models.call_log import CallLog
from app.schemas.promo_campaign import CampaignCreate, CampaignPreview, CampaignOut
from app.integrations.whatsapp import (
    list_approved_templates,
    list_all_templates,
    create_message_template,
    credentials_from_location,
)
from app.services.promo_campaign import (
    create_and_launch_campaign,
    launch_campaign_from_contacts,
    preview_audience,
)

router = APIRouter(prefix="/campaigns", tags=["campaigns"])


@router.get("/templates")
def list_wa_templates(
    location_id: uuid.UUID,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """List the location's APPROVED WhatsApp templates for the broadcast picker."""
    location = db.query(Location).filter(Location.id == location_id).first()
    if not location:
        raise HTTPException(status_code=404, detail={"message": "Location not found.", "code": "not_found"})
    waba_id = location.whatsapp_waba_id
    creds = credentials_from_location(location)
    token = creds.access_token if creds else None
    if not waba_id or not token:
        return {"connected": False, "templates": []}
    return {"connected": True, "templates": list_approved_templates(waba_id, token)}


@router.get("/templates/all")
def list_all_wa_templates(
    location_id: uuid.UUID,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """List every WhatsApp template for this location (any review status) — for the
    template management screen, so a submitted-but-pending template is visible."""
    location = db.query(Location).filter(Location.id == location_id).first()
    if not location:
        raise HTTPException(status_code=404, detail={"message": "Location not found.", "code": "not_found"})
    waba_id = location.whatsapp_waba_id
    creds = credentials_from_location(location)
    token = creds.access_token if creds else None
    if not waba_id or not token:
        return {"connected": False, "templates": []}
    return {"connected": True, "templates": list_all_templates(waba_id, token)}


class TemplateCreate(BaseModel):
    location_id: uuid.UUID
    name: str = Field(min_length=1)
    category: str  # MARKETING | UTILITY
    language: str = "en"
    body: str = Field(min_length=1)
    example_params: List[str] = []


@router.post("/templates", status_code=201)
def create_wa_template(
    body: TemplateCreate,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """Submit a new WhatsApp message template to Meta for review. Meta typically
    reviews within minutes to a few hours; once approved, it appears automatically
    in the campaign template picker — no further action needed here."""
    location = db.query(Location).filter(Location.id == body.location_id).first()
    if not location:
        raise HTTPException(status_code=404, detail={"message": "Location not found.", "code": "not_found"})
    waba_id = location.whatsapp_waba_id
    creds = credentials_from_location(location)
    token = creds.access_token if creds else None
    if not waba_id or not token:
        raise HTTPException(status_code=400, detail={"message": "WhatsApp is not connected for this location.", "code": "not_connected"})

    # Meta requires lowercase letters, digits, and underscores only.
    name = re.sub(r"[^a-z0-9_]", "_", body.name.strip().lower())
    if not name:
        raise HTTPException(status_code=400, detail={"message": "Template name is required.", "code": "invalid_name"})

    try:
        result = create_message_template(
            waba_id=waba_id,
            access_token=token,
            name=name,
            category=body.category,
            language=body.language or "en",
            body_text=body.body.strip(),
            example_params=body.example_params,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail={"message": str(e), "code": "meta_error"})

    return {
        "id": result.get("id"),
        "name": name,
        "status": (result.get("status") or "PENDING").upper(),
        "category": result.get("category", body.category.upper()),
    }


@router.get("", response_model=List[CampaignOut])
def list_campaigns(
    location_id: Optional[uuid.UUID] = None,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    q = db.query(PromoCampaign)
    if location_id:
        q = q.filter(PromoCampaign.location_id == location_id)
    return q.order_by(PromoCampaign.created_at.desc()).all()


@router.delete("/{campaign_id}", status_code=204)
def delete_campaign(
    campaign_id: uuid.UUID,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """Delete a campaign's history record. Any already-dispatched calls/messages are
    unaffected — this only removes the row from the campaign history list."""
    c = db.query(PromoCampaign).filter(PromoCampaign.id == campaign_id).first()
    if not c:
        raise HTTPException(status_code=404, detail={"message": "Campaign not found.", "code": "not_found"})
    db.delete(c)
    db.commit()


@router.get("/{campaign_id}/stats")
def campaign_stats(
    campaign_id: uuid.UUID,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """Delivery/outcome breakdown for one campaign, for the delivery dashboard."""
    c = db.query(PromoCampaign).filter(PromoCampaign.id == campaign_id).first()
    if not c:
        raise HTTPException(status_code=404, detail={"message": "Campaign not found.", "code": "not_found"})

    cid = str(campaign_id)
    out = {
        "id": cid, "name": c.name, "channel": c.channel.value,
        "total_targets": c.total_targets, "queued": c.messages_queued if c.channel == CampaignChannel.whatsapp else c.calls_queued,
    }

    if c.channel == CampaignChannel.whatsapp:
        rows = dict(
            db.query(WhatsAppMessage.status, func.count(WhatsAppMessage.id))
            .filter(WhatsAppMessage.campaign_id == cid, WhatsAppMessage.direction == WADirection.outbound)
            .group_by(WhatsAppMessage.status).all()
        )
        failed = rows.get(WAStatus.failed, 0)
        read = rows.get(WAStatus.read, 0)
        delivered = rows.get(WAStatus.delivered, 0) + read
        sent = sum(rows.values())  # every outbound broadcast row
        out["whatsapp"] = {"sent": sent, "delivered": delivered, "read": read, "failed": failed}
    else:
        rows = dict(
            db.query(CallLog.outcome, func.count(CallLog.id))
            .filter(CallLog.campaign_id == cid)
            .group_by(CallLog.outcome).all()
        )
        outcomes = {(k.value if k else "pending"): v for k, v in rows.items()}
        out["calls"] = {"total": sum(rows.values()), "outcomes": outcomes}

    return out


@router.post("/preview")
def preview_campaign(
    body: CampaignPreview,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """Return how many contacts this audience would reach, before launching."""
    count = preview_audience(
        db,
        location_id=body.location_id,
        audience=body.audience,
        tier=body.tier,
        expiring_days=body.expiring_days,
        lead_status=body.lead_status,
    )
    return {"count": count}


@router.post("", response_model=CampaignOut, status_code=201)
def create_campaign(
    body: CampaignCreate,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """Create a campaign and immediately queue staggered calls or WhatsApp messages."""
    if body.channel == CampaignChannel.call and not (body.message or "").strip():
        raise HTTPException(status_code=400, detail={"message": "An offer message is required for call campaigns.", "code": "no_message"})
    if body.channel == CampaignChannel.whatsapp and not body.wa_template:
        raise HTTPException(status_code=400, detail={"message": "Select a WhatsApp template.", "code": "no_template"})
    return create_and_launch_campaign(
        db=db,
        location_id=body.location_id,
        name=body.name,
        message=body.message or body.wa_template or "",
        audience=body.audience,
        channel=body.channel,
        tier=body.tier,
        expiring_days=body.expiring_days,
        lead_status=body.lead_status,
        wa_template=body.wa_template,
        wa_language=body.wa_language,
        wa_params=body.wa_params,
    )


@router.post("/import", response_model=CampaignOut, status_code=201)
async def import_campaign(
    location_id: uuid.UUID,
    name: str = Form(...),
    message: str = Form(""),
    channel: str = Form("call"),
    wa_template: str = Form(""),
    wa_language: str = Form("en"),
    wa_params: str = Form("[]"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """Launch a campaign (calls or WhatsApp) from an uploaded CSV or Excel contact list.
    Required column: phone. Optional: full_name (or name)."""
    try:
        chan = CampaignChannel(channel)
    except ValueError:
        chan = CampaignChannel.call
    if chan == CampaignChannel.call and not message.strip():
        raise HTTPException(status_code=400, detail={"message": "An offer message is required for call campaigns.", "code": "no_message"})
    if chan == CampaignChannel.whatsapp and not wa_template:
        raise HTTPException(status_code=400, detail={"message": "Select a WhatsApp template.", "code": "no_template"})
    try:
        params_list = json.loads(wa_params) if wa_params else []
    except Exception:
        params_list = []

    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(
            status_code=400,
            detail={"message": "Only CSV or Excel (.xlsx) files are accepted.", "code": "invalid_file"},
        )

    content = await file.read()
    rows: list[dict] = []
    try:
        if filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for excel_row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(excel_row)})
            wb.close()
        else:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = [{(k or "").strip().lower(): v for k, v in r.items()} for r in reader]
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail={"message": f"Could not parse file: {e}", "code": "parse_error"},
        )

    if not rows:
        raise HTTPException(
            status_code=400,
            detail={"message": "File is empty.", "code": "empty_file"},
        )

    return launch_campaign_from_contacts(
        db=db, location_id=location_id, name=name, message=message or wa_template or "", rows=rows,
        channel=chan, wa_template=wa_template or None, wa_language=wa_language or "en", wa_params=params_list,
    )

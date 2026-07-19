import csv
import io
import logging
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
import uuid

from app.core.database import get_db
from app.api.v1.auth import _get_current_user
from app.models.lead import Lead, LeadStatus
from app.models.customer import Language
from app.schemas.lead import LeadCreate, LeadUpdate, LeadOut
from app.services.lead import create_lead_and_trigger_outreach, bulk_create_leads

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/leads", tags=["leads"])


@router.get("", response_model=List[LeadOut])
def list_leads(
    location_id: Optional[uuid.UUID] = None,
    status: Optional[LeadStatus] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    q = db.query(Lead)
    if location_id:
        q = q.filter(Lead.location_id == location_id)
    if status:
        q = q.filter(Lead.status == status)
    if search:
        q = q.filter(
            Lead.full_name.ilike(f"%{search}%") | Lead.phone.ilike(f"%{search}%")
        )
    return q.order_by(Lead.created_at.desc()).all()


@router.post("", response_model=LeadOut, status_code=201)
def create_lead(
    body: LeadCreate,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    return create_lead_and_trigger_outreach(
        db=db,
        location_id=body.location_id,
        full_name=body.full_name,
        phone=body.phone,
        language=body.language,
        source=body.source,
        follow_up_date=body.follow_up_date,
    )


@router.post("/import", status_code=200)
async def import_leads_csv(
    location_id: uuid.UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """
    Import leads from a CSV or Excel (.xlsx) file.
    Required columns: full_name, phone
    Optional columns: language (en/hi/ta), source
    """
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(
            status_code=400,
            detail={"message": "Only CSV or Excel (.xlsx) files are accepted.", "code": "invalid_file"},
        )

    content = await file.read()
    rows: list[dict] = []

    try:
        if filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for excel_row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(excel_row)})
            wb.close()
        else:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = [{k.strip().lower(): v for k, v in r.items()} for r in reader]
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail={"message": f"Could not parse file: {e}", "code": "parse_error"},
        )

    if not rows:
        raise HTTPException(
            status_code=400,
            detail={"message": "File is empty.", "code": "empty_file"},
        )

    result = bulk_create_leads(db=db, location_id=location_id, rows=rows)
    return result


@router.get("/{lead_id}", response_model=LeadOut)
def get_lead(lead_id: uuid.UUID, db: Session = Depends(get_db), _=Depends(_get_current_user)):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail={"message": "Lead not found.", "code": "not_found"})
    return lead


@router.put("/{lead_id}", response_model=LeadOut)
def update_lead(
    lead_id: uuid.UUID,
    body: LeadUpdate,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail={"message": "Lead not found.", "code": "not_found"})
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(lead, field, value)
    db.commit()
    db.refresh(lead)
    return lead


@router.patch("/{lead_id}/stop", response_model=LeadOut)
def stop_lead_outreach(
    lead_id: uuid.UUID,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """Manually stop all outreach for a lead."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail={"message": "Lead not found.", "code": "not_found"})
    lead.wa_stopped = True
    lead.call_stopped = True
    db.commit()
    db.refresh(lead)
    return lead


@router.delete("/{lead_id}", status_code=204)
def delete_lead(
    lead_id: uuid.UUID,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail={"message": "Lead not found.", "code": "not_found"})
    from app.models.lead_sequence_step import LeadSequenceStep
    db.query(LeadSequenceStep).filter(LeadSequenceStep.lead_id == lead_id).delete()
    db.delete(lead)
    db.commit()


@router.patch("/{lead_id}/convert", response_model=LeadOut)
def convert_lead(
    lead_id: uuid.UUID,
    db: Session = Depends(get_db),
    _=Depends(_get_current_user),
):
    """Mark a lead as converted (became a customer) and stop all outreach."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail={"message": "Lead not found.", "code": "not_found"})
    lead.status = LeadStatus.converted
    lead.wa_stopped = True
    lead.call_stopped = True
    db.commit()
    db.refresh(lead)
    return lead
